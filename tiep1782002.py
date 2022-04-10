from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from PIL import Image,ImageFilter
import pytesseract as pt
from pytesseract import image_to_string
from shutil import copyfile
import time , requests,PIL
import openpyxl
import random,os,re,io,sys
from multiprocessing import Pool
pt.pytesseract.tesseract_cmd = r'c:\Program Files\Tesseract-OCR\Tesseract.exe'
tang=0
def update_value_excel(filename, cellname, value):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet1']
    Sheet1[cellname].value = value
    wb.close()
    wb.save(filename)

def initDriver(filepath):

	options = webdriver.ChromeOptions()
	options.add_argument(r'user-data-dir=c:\withpython\\'+ filepath) 
	
	prefs = {
	"profile.managed_default_content_settings.images": 2
	}
	options.add_experimental_option("prefs", prefs)
	
	browser = webdriver.Chrome(executable_path=r'c:\withpython\chromedriver0.exe', chrome_options=options)
	return browser 

def suly(img):
    row,col = img.size
    data=[] #r,g,b,i,j
    pixels=img.load()

    xtam=int(row/2)
    ytam=int(col/2)
    #tim start
    chaystart = 0
    #chạy đến captcha 
    while (True):
        start = xtam-10
        end = xtam+10
        dk= True
        while(start < end):
            #nếu là màu nền
            dk1=bool(218 == pixels[start,chaystart][0] and 112 == pixels[start,chaystart][1] and 214 == pixels[start,chaystart][2] )
            dk2=bool(165 == pixels[start,chaystart][0] and 42 == pixels[start,chaystart][1] and 42 == pixels[start,chaystart][2] )
            dk3=bool(128 == pixels[start,chaystart][0] and 128 == pixels[start,chaystart][1] and 128 == pixels[start,chaystart][2] )
            dk4=bool(128 == pixels[start,chaystart][0] and 0 == pixels[start,chaystart][1] and 128 == pixels[start,chaystart][2] )
            dk5=bool(0 == pixels[start,chaystart][0] and 0 == pixels[start,chaystart][1] and 0 == pixels[start,chaystart][2] )
            dk *= bool( dk1 + dk2 + dk3 + dk4 + dk5 )
            start +=1
            if(dk == False):
                #pixels[start,chaystart]=(0,0,0)
                break
            
        if(dk == True):
            break
        chaystart+=1

    #tim end
    chayend = col-1
    #chạy đến captcha từ dưới lên
    while (True):
        start = xtam-10
        end = xtam+10
        dk= True
        while(start < end):
            #nếu là màu nền
            dk1=bool(218 == pixels[start,chayend][0] and 112 == pixels[start,chayend][1] and 214 == pixels[start,chayend][2] )
            dk2=bool(165 == pixels[start,chayend][0] and 42 == pixels[start,chayend][1] and 42 == pixels[start,chayend][2] )
            dk3=bool(128 == pixels[start,chayend][0] and 128 == pixels[start,chayend][1] and 128 == pixels[start,chayend][2] )
            dk4=bool(128 == pixels[start,chayend][0] and 0 == pixels[start,chayend][1] and 128 == pixels[start,chayend][2] )
            dk5=bool(0 == pixels[start,chayend][0] and 0 == pixels[start,chayend][1] and 0 == pixels[start,chayend][2] )
            dk *= bool( dk1 + dk2 + dk3 + dk4 + dk5 )
            start +=1
            #pixels[start,chayend]=(0,0,0)
            if(dk == False):
                break
            
        if(dk == True):
            break
        chayend-=1


    # tim x max
    chay=0
    so_lan_true = 0
    while (so_lan_true <= 40 ):
        #cho duyệt 1 cột pixel từ tâm ra ngoài , nếu cột này có các màu bằng nhau thì lấy làm mốc xmax
        start = chaystart+1
        end = chayend-1
        dk= True
        while(start < end):
            dk *= bool( pixels[xtam+chay,ytam][0] ==  pixels[xtam+chay,start][0] and pixels[xtam+chay,ytam][1] ==  pixels[xtam+chay,start][1] and pixels[xtam+chay,ytam][2] ==  pixels[xtam+chay,start][2])
            start +=1
            if(dk == False):
                so_lan_true = 0
                break
            
        if(dk == True):
            so_lan_true +=1
        chay+=1
    xmax = xtam+chay-40

    #tim x min
    chay=0
    so_lan_true = 0
    while (so_lan_true <= 40):
        start = chaystart
        end = chayend
        dk= True
        while(start < end):
            dk *= bool( pixels[xtam+chay,ytam][0] ==  pixels[xtam+chay,start][0] and pixels[xtam+chay,ytam][1] ==  pixels[xtam+chay,start][1] and pixels[xtam+chay,ytam][2] ==  pixels[xtam+chay,start][2])
            start +=1
            if(dk == False):
                so_lan_true = 0
                break
            
        if(dk == True):
            so_lan_true +=1
        chay-=1
    xmin = xtam+chay+40
    ##############################################################################################################################################

    #sử lí ảnh 
    img = img.crop((xmin,chaystart, xmax,chayend))
    img_data = img.getdata()

    lst=[]
    for i in img_data:
        #nếu là màu chữ
        dk1=bool(0 == i[0] and 0 == i[1] and 255 == i[2] )
        dk2=bool(255 == i[0] and 165 == i[1] and 0 == i[2] )
        dk3=bool(0 == i[0] and 128 == i[1] and 0 == i[2] )
        dk4=bool(0 == i[0] and 255 == i[1] and 255 == i[2] )
        dk5=bool(255 == i[0] and 215 == i[1] and 0 == i[2] )
        dk6=bool(255 == i[0] and 255 == i[1] and 255 == i[2] )
        dk7=bool(255 == i[0] and 0 == i[1] and 0 == i[2] )
        
        dk = bool( dk1 + dk2 + dk3 + dk4 + dk5 + dk6 + dk7 )
        if (dk):
            lst.append(0)
        
        else:
            lst.append(255)
        
    new_img = Image.new("L", img.size)
    new_img.putdata(lst)
    return image_to_string(new_img,lang="vie")

def capcha(xy):
	#--------------------------------------------------------------------------------------
	global tang
	print('------BAT DAU GIAI CAPTCHA--------')
	driver.switch_to.window (driver.window_handles [0])
	driver.set_window_size(1043, 1200)
	time.sleep(4)
	driver.get_screenshot_as_file("d:\capcha"+taikhoanx+".png")
	image =  Image.open("d:\capcha"+taikhoanx+".png")
	
	key_captcha = suly(image)
	print(key_captcha)
	if os.path.exists("d:\capcha"+taikhoanx+".png"):
		os.remove("d:\capcha"+taikhoanx+".png")

	#su ly key_captcha
	kt=key_captcha[len(key_captcha)-1]
	print('___________',kt)
	key_captcha = re.sub( kt, '', key_captcha)
	key_captcha = re.sub('\?|\'|\§|\ |\:|\\n','', key_captcha)
	print(key_captcha)
	#lay dap an
	dapan=0
	j=0
	if(key_captcha.find('bala') > -1 ):
		dapan=3
	if(key_captcha.find('hai') > -1 and key_captcha.find('la') > -1 ):
		dapan=2
	if(key_captcha.find('Xe0tOcobao') > -1  ):
		dapan=4
	if(key_captcha.find('S6hai') > -1  ):
		dapan=2
	if(key_captcha.find('diénthoai') > -1  ):
		dapan=10
	if(key_captcha.find('8-5') > -1  ):
		dapan=3
	if(key_captcha.find('2bantay') > -1  ):
		dapan=10
	if(key_captcha.find('ONHO') > -1  ):
		dapan=5
	if(key_captcha.find('tamla') > -1  ):
		dapan=8
	if(key_captcha.find('Convit') > -1  ):
		dapan=2
	if(key_captcha.find('conga') > -1  ):
		dapan=2
	if(key_captcha.find('con') > -1 and key_captcha.find('ga') > -1 ):
		dapan=2
	if(key_captcha.find('chin') > -1 and key_captcha.find('la') > -1 ):
		dapan=9
	if(key_captcha.find('mot') > -1 and key_captcha.find('la') > -1 ):
		dapan=1
	if(key_captcha.find('6mudi') > -1 and key_captcha.find('la') > -1 ):
		dapan=10
	
	if(key_captcha.find('sao') > -1 and key_captcha.find('bao') > -1 ):
		dapan=5
	if(key_captcha.find('ébaonhiéumat') > -1  and key_captcha.find('ID') > -1  ):
		dapan=1
	if(key_captcha.find('NAM') > -1  ):
		dapan=7
	if(key_captcha.find('NE12') > -1  ):
		dapan=8
	if(key_captcha.find('VIEW') > -1  ):
		dapan=6
	if(key_captcha.find('LIKE') > -1  ):
		dapan=6
	if(key_captcha.find('9+1') > -1  ):
		dapan=10
	if(key_captcha.find('9') > -1 and key_captcha.find('+') > -1 and key_captcha.find('1') > -1  ):
		dapan=10
	if(key_captcha.find('2') > -1 and key_captcha.find('+') > -1 and key_captcha.find('3') > -1  ):
		dapan=5
	if(key_captcha.find('8') > -1 and key_captcha.find('-') > -1 and key_captcha.find('5') > -1  ):
		dapan=3
	
	if(key_captcha.find('tamgi') > -1  ):
		dapan=3
	
	if(key_captcha.find('bonla') > -1  ):
		dapan=4
	
	if(key_captcha.find('choco') > -1  ):
		dapan=4
	
	if(key_captcha.find('S6mu') > -1  ):
		dapan=10
	
	if(key_captcha.find('onngu') > -1 and  key_captcha.find('chan') > -1):
		dapan=2
	
	if(key_captcha.find('Concd') > -1  ):
		dapan=2
	
	if(key_captcha.find('Conco') > -1  ):
		dapan=2
	
	if(key_captcha.find('namla') > -1  ):
		dapan=5
	if(key_captcha.find('saula') > -1  ):
		dapan=6
	if(key_captcha.find('bayla') > -1  ):
		dapan=7
	
	if(key_captcha.find('ConMe') > -1  ):
		dapan=1
	
	if(key_captcha.find('Xe0t6') > -1  ):
		dapan=4
	
	if(key_captcha.find('Hinh') > -1 and key_captcha.find('nhat') > -1  ):
		dapan=4
	
	if(key_captcha.find('2+3') > -1  ):
		dapan=5
	
	if(key_captcha.find('Xeda') > -1  ):
		dapan=2
	
	if(key_captcha.find('Hinhvu') > -1  ):
		dapan=4
	
	if(key_captcha.find('Convo') > -1  ):
		dapan=1
	if(key_captcha.find('conga') > -1 or key_captcha.find('Conga') > -1  ):
		dapan=2
	if(key_captcha.find('con') > -1 and key_captcha.find('voi') > -1  ):
		dapan=1
	if(key_captcha.find('mat') > -1 and key_captcha.find('ang') > -1  ):
		dapan=1
	if(key_captcha.find('b6nla') > -1  ):
		dapan=4
	if(key_captcha.find('lialamt') > -1 and key_captcha.find('canh') > -1  ):
		dapan=3
	if(key_captcha.find('CLUB') > -1  ):
		dapan=6
	if(key_captcha.find('Hinh') > -1 and key_captcha.find('vudng') > -1  ):
		dapan=4
	if(key_captcha.find('onng') > -1 and key_captcha.find('mat') > -1  ):
		dapan=2
	
	if(key_captcha.find('ncho') > -1  ):
		dapan=4
	if(key_captcha.find('ConIon') > -1  ):
		dapan=4
	if(key_captcha.find('Xe6t') > -1  ):
		dapan=4
	if(key_captcha.find('Hinhvuéng') > -1  ):
		dapan=4
	if(key_captcha.find('1ti') > -1  ):
		dapan=1
	if(key_captcha.find('1tu') > -1 ):
		dapan=1
	if(key_captcha.find('Laco') > -1  ):
		dapan=1
	if(key_captcha.find('Nam') > -1  ):
		dapan=1
	
	if(key_captcha.find('BOOK') > -1  ):
		dapan=8
	
	if(key_captcha.find('Conlon') > -1  ):
		dapan=4
	
	if(key_captcha.find('bdnla') > -1  ):
		dapan=4
	
	if(key_captcha.find('ConIdn') > -1  ):
		dapan=4
	
	if(key_captcha.find('Xem') > -1  ):
		dapan=2
	
	if(key_captcha.find('2ban') > -1  ):
		dapan=10
	
	if(key_captcha.find('vuéng') > -1  ):
		dapan=4
	
	if(key_captcha.find('tay') > -1 and key_captcha.find('cai') > -1 ):
		dapan=2
	
	if(key_captcha.find('métla') > -1  ):
		dapan=1
	
	if(key_captcha.find('baonhiéus6') > -1  ):
		dapan=10
	
	if(key_captcha.find('matte)!') > -1  ):
		dapan=1
	
	if(key_captcha.find('1') > -1 and key_captcha.find('ngay') > -1  ):
		dapan=7
	
	if(key_captcha.find('1bant') > -1  ):
		dapan=5
	
	if(key_captcha.find('onIgn') > -1  ):
		dapan=2
	
	if(key_captcha.find('ngon') > -1 and key_captcha.find('ai') > -1  ):
		dapan=2
	
	
	if(key_captcha.find('S6nam') > -1  ):
		dapan=5
	if(key_captcha.find('Cobaonhiéumat') > -1  ):
		dapan=1
	'''
	if(key_captcha.find('2bantaycébao') > -1  ):
		dapan=
	if(key_captcha.find('2bantaycébao') > -1  ):
		dapan=
	wb = openpyxl.load_workbook("captcha"+str(index)+".xlsx"filename)
    Sheet1 = wb['Sheet1']
    wb.close()
    return Sheet1[cellname].value
	'''
	if(dapan == 0 ):
		mo = openpyxl.load_workbook("captcha"+str(index)+".xlsx")
		wb = mo['Sheet1']
		while ( wb['a'+str(j+2)].value != None):
			
			if(key_captcha == wb['a'+str(j+2)].value):
				dapan = wb['b'+str(j+2)].value
				break
			j+=1
		mo.close()

	if (dapan == 0):
		
		print('-------NHAP DAP AN CAPCHA-------')
		
		'''
		driver.get_screenshot_as_file("nnnnnnnnnnnnncapcha"+taikhoanx+".png")
		driver.quit()
		time.sleep(1)
		tang += 1
		minimain(index)
		return False
		'''



		dk1=True
		while(dk1):
			try:
				dapan =int(input())
			except:
				print ("###################################################################")
				dk1=True
			dk1=False
			print("___",dapan,"_____")
			if( dapan != 10 and  dapan != 9 and  dapan != 8 and  dapan != 7 and  dapan != 6 and  dapan != 5 and  dapan != 4 and  dapan != 3 and  dapan != 2 and  dapan != 1 ):
				print("KHONG THOA MAN , NHAP LAI ")
				dk1=True
		update_value_excel("captcha"+str(index)+".xlsx", 'a'+str(j+2), key_captcha)
		update_value_excel("captcha"+str(index)+".xlsx", 'b'+str(j+2), dapan)
	print('-------DAP AN CUA CAPTCHA LA---------: ', dapan)
	
	
	#DA CO DAP AN , BAT DAU RECAPTCHA
	if (dapan == 1 ):
		driver.find_element_by_tag_name('select').send_keys(str(dapan))
		time.sleep(1)
	driver.find_element_by_tag_name('select').send_keys(str(dapan))
	time.sleep(1.5)
	if(xy == 1 ):
		driver.execute_script("document.getElementsByTagName('button')[1].click()")
	if (xy == 2 ):
		driver.execute_script("document.getElementsByTagName('button')[3].click()")
	time.sleep(1)
	if(xy ==1 ):
		xetcapcha()
	#else:
	#	xetcapchainstar()

def xetcapcha():
	global tang
	#print(" ------------ĐANG XÉT CAPTCHA ------------------")
	driver.switch_to.window (driver.window_handles [0])
	driver.set_window_size(1043, 1200)
	time.sleep(1)
	driver.get_screenshot_as_file("d:\capcha"+taikhoanx+".png")
	im2 =  Image.open("d:\capcha"+taikhoanx+".png")
	text = pt.image_to_string(im2)
	#print(text)
	if os.path.exists("d:\capcha"+taikhoanx+".png"):
		os.remove("d:\capcha"+taikhoanx+".png")
	x = text.find('Chon cau')
	y=text.find('icon')
	if(x != -1 or y  != -1 or text.find('tra loi') != -1):
		#print('co captcha huhu')
		
		dk2 = capcha(1)
		if (dk2 == False):
			time.sleep(0.2)
			return False
			return False

	elif(text.find('Home') != -1 or text.find('Menu') > -1 or text.find('avatar') != -1  or text.find('Home') >0 or text.find('Chon tai khoan') != -1 or text.find('quay lai sau 10 giay') != -1 or text.find('Kiém') != -1 or text.find('Chon tai') != -1):
		try:
			print(text.find('Home'),"-----------","KHÔNG CÓ BẢNG THÔNG BÁO HIỆN LÊN =================================")
		finally:
			return True
		return True
	elif(text.find('Thong bao') != -1 ):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	elif(text.find('Network') != -1 ):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	elif(text.find('Error') != -1):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	elif(text.find('Network') != -1 ):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	elif(text.find('Error') != -1):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	
	
	elif(text.find('Ban da lam qua 100 jobs') != -1):
		#driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		print("=============================CHUYỂN NICK NÀO ============================")
		driver.quit()
		time.sleep(1)
		tang += 1
		minimain(index)
	elif(text.find('Thanh cong') != -1 or text.find('LAG') != -1 ):
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
		return True
	else:
		
		try:
			cten = driver.find_element_by_id('swal2-content')
			print("cten dax duoc thiet lap --------------------------------------------------------------------------------")
		except:
			print("capTcha dac biet -------------------------------------")
			capcha( 1)
		try:
			print(cten.text)
			print("cten dax duoc thiet lap ============================================================================")
		except:
			xetcapcha()
		try:
			cten.text.find('Đã gửi thông tin lên hệ thống xét duyệt')
		except:
			xetcapcha()
		if(cten.text.find('Đã gửi thông tin lên hệ thống xét duyệt') != -1 ):
			binh_thuong()
		elif(cten.text.find('Hệ thống jobs đang trong quá trình xử lý chặn auto và phân phối jobs') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			i=0
			while(i < 300):
				print("ĐANG ĐỢI ",300-i,"===========================================")
				time.sleep(1)
				i+=1
			return True
	
		elif(text.find('Network') != -1 ):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(text.find('Error') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('quá nhanh') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			try:
				driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
				time.sleep(0.7)
				driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			except:
				to_cao()
				time.sleep(0.2)
			return True
		elif(cten.text.find('Bạn đã làm quá 100 jobs') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			print("=============================CHUYỂN NICK NÀO ============================")
			driver.quit()
			time.sleep(1)
			tang += 1
			minimain(index)
		elif(cten.text.find('quá 3 lần') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			print("=============================CHUYỂN NICK NÀO ============================")
			driver.quit()
			time.sleep(1)
			tang += 1
			minimain(index)
		elif(cten.text.find('LAG') != -1 ):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Bài viết đã hết hạn hoàn thành') != -1 ):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Vui lòng thử lại sau ít phút, hoặc liên hệ admin để hỗ trợ') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Vui lòng cập nhật phiên bản mới nhất để làm việc') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Để đảm bảo công bằng cho mọi người') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Đã gửi báo cáo lên hệ thống') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Tài khoản của bạn KHÔNG sẵn sàng làm jobs like page') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		elif(cten.text.find('Bài viết này đã bị ẩn khi bị mọi người báo cáo lỗi quá nhiều') != -1):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
			return True
		else:
			print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@\n@@@@@@@@@@@@@@@@@@@@@@@@@@@")
			#input()
			print("=============================LÀM LẠI THÔI============================")
			driver.quit()
			minimain(index)
	#time.sleep(3)
def to_cao():
	
	time.sleep(random.randint(3,4))
	try:
		to=driver.find_elements_by_tag_name("i")
	except:
		return False
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem ==  22):
		driver.execute_script("document.getElementsByTagName('i')[6].click()")#chon loai to cao
	elif(dem == 23):
		driver.execute_script("document.getElementsByTagName('i')[7].click()")#chon loai to cao
	elif(dem == 19):
		driver.execute_script("document.getElementsByTagName('i')[4].click()")#chon loai to cao
	else:
		try:
			driver.execute_script("document.getElementsByTagName('i')[5].click()")#chon loai to cao
		except:
			return False
	time.sleep(random.randint(3,4))
	try:
		driver.execute_script("document.getElementsByClassName('btn btn-primary btn-sm form-control mt-3')[0].click()")
	except:
		return False
	time.sleep(2)
	try:
		driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
	except:
		return False

def binh_thuong():
	driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok
	time.sleep(random.randint(2,5))
	xetcapcha()

def like_fage():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo link job
	except:
		to_cao()
	time.sleep(random.randint(1,6))

	time.sleep(random.randint(4,9))
	driver.switch_to.window (driver.window_handles [1])
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(6,15))
	driver.execute_script("window.scrollTo(0, 0);")
	try:
		driver.execute_script("document.getElementsByClassName('ct')[0].click()")#nhanlikefage
	except:
		huy_job()
	time.sleep(random.randint(4,9))
	driver.close()
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
		

	return True	
def theo_doi():
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True

	xetcapcha()
	driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job

	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])

	leng = driver.find_elements_by_tag_name("input")
	if(len(leng) == 6):
		driver.execute_script("document.getElementsByTagName('input')[5].click()")
	try:
		driver.find_element_by_link_text("Theo dõi").click()
	except:
		time.sleep(0.2)
	
	time.sleep(random.randint(1,2))
	
	driver.switch_to.window (driver.window_handles [1])
	driver.close()
	
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True	
def like_bai():
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True

	xetcapcha()
	driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job

	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])

	leng = driver.find_elements_by_tag_name("input")
	if(len(leng) == 6):
		driver.execute_script("document.getElementsByTagName('input')[5].click()")
	
	time.sleep(random.randint(1,2))
	driver.find_element_by_link_text("Bày tỏ cảm xúc").click()
	huy = 0
	time.sleep(random.randint(1,2))
	
	
	nhan=driver.find_elements_by_tag_name('li')
	
	
	
	nhan[0].click()
	
	time.sleep(random.randint(4,9))
	
	driver.switch_to.window (driver.window_handles [1])
	driver.close()
	
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def thuong_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		 
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[2].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(4,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def love_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		 
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[1].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(4,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def phanno_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[6].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(4,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def buon_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		 
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[5].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(4,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def tang_comment():
	driver.switch_to.window (driver.window_handles [0])
	driver.execute_script("document.getElementsByTagName('u')[0].click()")#nhan copy
	
	thecomment = driver.find_elements_by_tag_name("span")
	comment = thecomment[7].text
	mbasic= driver.find_elements_by_tag_name("h6")
	for mb in mbasic:
		mb.text.lower() == "mbasic"
		mb.click()
	driver.switch_to.window (driver.window_handles [1])
	time.sleep(1)
	driver.find_element_by_name("comment_text").send_keys(comment)
	time.sleep(random.randint(3,4))
	driver.find_element_by_xpath("//option[text()='Bình luận']").click()
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(1)
	for mb in mbasic:
		mb.text.lower() == "hoàn thành"
		mb.click()
	
	xetcapcha()
	
	return True
def lamviec ():
	time.sleep(random.randint(4,9))
	
	dklm = xetcapcha()
	if (dklm == False): return False

	try:
		element = WebDriverWait(driver, 100).until(
			EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div[1]/div[2]/div/div[1]/div[2]/span/div[1]"))
		)
	except:
		try:
			print("---------------------đang thử tìm mục trở lại-------------------------")
			dk4 = driver.find_elements_by_link_text("Chi tiết")
			if (dk4 != None):
				time.sleep(random.randint(2,4))
				print("co muc tro lai")
				driver.execute_script("document.getElementsByTagName('i')[0].click()")# nhan tro lai
				try:
					driver.find_elements_by_link_text("Chọn kênh kiếm tiền")
					time.sleep(random.randint(2,4))
					driver.execute_script("document.getElementsByClassName('card shadow-400 h-100 mb-3 hand bg-gradient-1')[0].click()")
				except:
					print("nnnn")
		except:
			dklm = xetcapcha()
			if (dklm == False): return False
		finally:
			lamviec()
	#===================================================================================================================================================================
	try:
		driver.execute_script("document.getElementsByClassName('card mb-2 hand')[0].click()")#nhan job
	except:
		dklm = xetcapcha()
		if (dklm == False): return False
		lamviec()
	time.sleep(random.randint(4,9))
	try:
		xacnhan = driver.find_elements_by_tag_name('span')
		print(xacnhan[5].text)
	except:to_cao()
	time.sleep(random.randint(3,4))
	

	if (xacnhan[5].text == 'TĂNG LIKE CHO FANPAGE'):
		#to_cao()
		like_fage()
	elif(xacnhan[5].text == 'TĂNG LIKE CHO BÀI VIẾT'):
		like_bai()
	elif(xacnhan[5].text == 'TĂNG THƯƠNG THƯƠNG CHO BÀI VIẾT'):
		thuong_bai()
	elif(xacnhan[5].text == 'TĂNG LOVE CHO BÀI VIẾT'):
		love_bai()
	elif(xacnhan[5].text == 'TĂNG HAHA CHO BÀI VIẾT'):
		haha_bai()
	elif(xacnhan[5].text == 'TĂNG WOW CHO BÀI VIẾT'):
		wow_bai()
	elif(xacnhan[5].text == 'TĂNG BUỒN CHO BÀI VIẾT'):
		buon_bai()
	elif(xacnhan[5].text == 'TĂNG PHẪN NỘ CHO BÀI VIẾT'):
		phanno_bai()
	elif(xacnhan[5].text == 'TĂNG COMMENT CHO BÀI VIẾT'):
		tang_comment()
	elif(xacnhan[5].text == 'TĂNG LƯỢT THEO DÕI'):
		theo_doi()
	else:
		print("ngoại lệ : không biết đây là job gì ")
		
		dklm = xetcapcha()
		if (dklm == False): return False
	return dklm
def huy_job():
	try:
		driver.switch_to.window (driver.window_handles [1])
	except:
		return True
	driver.close()
	#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ chua toi uu @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	to_cao()
def haha_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[3].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(3,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def wow_bai():
	time.sleep(random.randint(3,4))
	to=driver.find_elements_by_tag_name("i")
	dem = 0
	for i in to :
		dem+=1
	print ('co ', dem, " classssssssssssssssssssssssssssssss-------------------------------------------------------------------")
	if (dem !=  22):							
		to_cao()
		return True
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[4].click()")#mo job
	except:
		to_cao()
	time.sleep(random.randint(1,9))
	driver.switch_to.window (driver.window_handles [1])
	try:
		leng = driver.find_elements_by_tag_name("input")
		if(len(leng) == 6):
			driver.execute_script("document.getElementsByTagName('input')[5].click()")
	except:
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
	time.sleep(random.randint(2,3))
	driver.execute_script("window.scrollTo(0, 0);")
	time.sleep(random.randint(1,2))
	
	try:
		thuong = driver.find_elements_by_link_text('Bày tỏ cảm xúc')#nhan thuong bai
		
		huy = 0
	except:
		huy=1
		huy_job()
	if (thuong == None):
		if(huy == 0 ):huy_job()
	try:
		thuong[0].click()
	except:
		if(huy == 0 ):huy_job()
	try:
		nhan=driver.find_elements_by_tag_name('li')
	except:
		if(huy == 0 ):huy_job()
	if (nhan == None):
		if(huy == 0 ):huy_job()
	try:
		nhan[4].click()
	except:
		if(huy == 0 ):huy_job()
	time.sleep(random.randint(4,9))
	try:
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, (document.body.scrollHeight)/2);")
		time.sleep(random.randint(1,2))
		driver.execute_script("window.scrollTo(0, 0);")
		time.sleep(random.randint(1,2))
		driver.switch_to.window (driver.window_handles [1])
		driver.close()
	except:
		driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(3,4))
	try:
		xetcapcha()
		driver.execute_script("document.getElementsByClassName('font-bold font-18')[6].click()")#hoan thanh
	except:
		to_cao()
	time.sleep(random.randint(3,4))
	
	xetcapcha()
	
	return True
def nuoi_fb():
	global window_nuoi
	driver.switch_to.window (golike)
	driver.execute_script("window.open('https://mbasic.facebook.com/','_blank')")
	time.sleep(1)
	driver.switch_to.window (driver.window_handles [1]) 
	time.sleep(0.5)
	window_nuoi =driver.current_window_handle
	print("cua so nuoi fb:",window_nuoi)
	driver.execute_script("javascript:setInterval(function(){window.scrollBy(0,(Math.floor((Math.random() * 1) + 0)+0.2)*window.innerHeight);}, Math.floor((Math.random() * 20000) + 10000));")
	time.sleep(0.5)
	driver.switch_to.window (golike)
def lam_fb():
	
	time.sleep(random.randint(3,4))
	driver.execute_script("document.getElementsByClassName('card shadow-400 h-100 mb-3 hand bg-gradient-1')[0].click()")
	xet = True
	while (xet):
		xet = lamviec()	
		print('lam viec tiep :',xet)
	return False
	'''
def lam_viec_instar():
	time.sleep(1)
	driver.execute_script("document.getElementsByClassName('btn btn-outline-light')[0].click()")#nhan job
	time.sleep(1)
	xacnhan = driver.find_elements_by_tag_name('span')

	print(xacnhan[3].text)
	time.sleep(random.randint(1,6))

	time.sleep(2)
	driver.execute_script("document.getElementsByClassName('btn bg-button-1 px-0 btn-block')[0].click()")#mo instar
	if (xacnhan[3].text == 'TĂNG LƯỢT THEO DÕI'):
		theo_doi_instar()
	elif(xacnhan[3].text == 'TĂNG LIKE CHO BÀI VIẾT' or xacnhan[3].text.find("TĂNG LIKE CHO BÀI VIẾT") >0):
		tim_bai_instar()
	else: input()
def lam_instar():
	
	time.sleep(random.randint(4,9))
	driver.execute_script("document.getElementsByClassName('card shadow-400 h-100 mb-3 hand bg-gradient-7')[0].click()")
	xet = True
	while (xet):
		xet = lam_viec_instar()
		print('lam viec tiep :',xet)
def theo_doi_instar():
	time.sleep(random.randint(4,9))
	driver.switch_to.window (driver.window_handles [1])
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,5))
	driver.execute_script("window.scrollTo(0, 0);")

	driver.execute_script("document.getElementsByTagName('button')[0].click()")#nhan follow
	time.sleep(random.randint(4,9))
	driver.close()
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(5,8))

	#hoan thanh
	button = driver.find_elements_by_tag_name('button')
	so=0
	for b in button:
		if (b.text == 'Hoàn thành'):
			hehe=so
		so+=1
	driver.execute_script("document.getElementsByTagName('button')["+str(hehe)+"].click()")
	xetcapchainstar()

	return True
	
def tim_bai_instar():
	time.sleep(random.randint(4,9))
	driver.switch_to.window (driver.window_handles [1])
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	time.sleep(random.randint(2,5))
	driver.execute_script("window.scrollTo(0, 0);")

	driver.execute_script("document.getElementsByClassName('QBdPU ')[1].click()")#nhan tim 
	time.sleep(random.randint(4,9))	
	driver.close()
	driver.switch_to.window (driver.window_handles [0])
	time.sleep(random.randint(5,8))

	#hoan thanh
	button = driver.find_elements_by_tag_name('button')
	so=0
	for b in button:
		if (b.text == 'Hoàn thành'):
			hehe=so
		so+=1
	driver.execute_script("document.getElementsByTagName('button')["+str(hehe)+"].click()")
	xetcapchainstar()
	return True
def xetcapchainstar():
	driver.set_window_size(1043, 1200)
	time.sleep(5)
	driver.get_screenshot_as_file("capcha"+taikhoanx+".png")
	im2 =  Image.open("capcha"+taikhoanx+".png")
	text = pt.image_to_string(im2)
	print(text)
	
	x = text.find('Chon cau')
	
	if(x != -1):
		print('co captcha huhu')
		
		capcha(im2,2)
	elif(text.find('Home') != -1 or text.find('Menu') != -1):
		return True
	else:
		cten = driver.find_element_by_id('swal2-content')
		print(cten.text)
		if(cten.text.find("thành công") != -1 ):
			driver.execute_script("document.getElementsByClassName('swal2-confirm swal2-styled')[0].click()")#nhan ok 
		else:input()
	'''
def minimain (x):
	taikhoan =  ("tiep1782002","tieple247","ductieple40","tiepvivu","tiepgolike09l","tiepgolike08l","tiepgolike06l","tiepgolike03","tiepgolike02","tiepgolike05l","tiepgolike04l")
	global driver,taikhoanx,tang,index,golike
	
	index=x
	while (True):
		if((index+tang*2) > len(taikhoan)-1):
			tang=0
		print(index+tang*2)
		taikhoanx = taikhoan[index+tang*2]
		if os.path.exists("d:\capcha"+taikhoanx+".png"):
			os.remove("d:\capcha"+taikhoanx+".png")
		driver = initDriver(taikhoanx)
		driver.get('https://golike.mobi/')
		time.sleep(random.randint(3,4))
        
		driver.execute_script("document.getElementsByClassName('font-13')[1].click()")
		time.sleep(2)
		golike =  driver.current_window_handle
		lam_fb()
#def aa ():
#	Pool(2).map(minimain, (0,2))
if __name__ == '__main__':
	
	#aa()
	minimain(0)
    